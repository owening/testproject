from openpyxl import load_workbook

wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['my_sheet']
print(sheet_ranges['A23'].value)

for i in range(1,31):
    print(sheet_ranges.cell(column=1, row=i).value)